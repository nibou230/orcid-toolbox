import streamlit as st
import re
from io import BytesIO
import pandas as pd
from src.orcid_data import fetch_orcid_data, format_timestamp
from src.references_matching import extract_and_process_references, prepare_orcid_works, match_references_to_orcid
from src.overton_data import get_overton_set_url
from src.format_citations import get_citations
import importlib.util
import gettext
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv
# TODO: look into using https://docs.python.org/3/library/concurrent.futures.html for parallel
# data fetching and processing.

# Settings
overton_enabled = True

# Load environment variables from .env file
load_dotenv()
overton_key = os.getenv("OVERTON_KEY")

# Set locale from Streamlit context if available, otherwise default to fr
default_locale = "fr"
if hasattr(st.context, "locale"):
    browser_locale = st.context.locale
    if isinstance(browser_locale, str) and browser_locale.startswith("fr"):
        default_locale = "fr"
    elif isinstance(browser_locale, str) and browser_locale.startswith("en"):
        default_locale = "en"

if "locale" not in st.session_state:
    st.session_state.locale = default_locale

# Keep locale in sync with sidebar selector without forcing manual reruns.
if "locale_picker" in st.session_state and st.session_state.locale != st.session_state.locale_picker:
    st.session_state.locale = st.session_state.locale_picker

# Set up gettext translations
_ = gettext.translation('messages', localedir='loc', languages=[st.session_state.locale], fallback=True).gettext

def reset_session_state():
    for key in list(st.session_state.keys()):
        st.session_state.pop(key)

st.set_page_config(page_title=_("app-title"), page_icon=":toolbox:", layout="wide", initial_sidebar_state="expanded")

# Human readable labels for work types
type_labels = {
                            "book": _("book"),
                            "book-chapter": _("book-chapter"),
                            "edited-book": _("edited-book"),
                            "conference-paper": _("conference-paper"),
                            "conference-output": _("conference-output"),
                            "conference-presentation": _("conference-presentation"),
                            "conference-poster": _("conference-poster"),
                            "conference-proceedings": _("conference-proceedings"),
                            "conference-abstract": _("conference-abstract"),
                            "journal-article": _("journal-article"),
                            "preprint": _("preprint"),
                            "dissertation-thesis": _("dissertation-thesis"),
                            "working-paper": _("working-paper"),
                            "other": _("other"),
                            "annotation": _("annotation"),
                            "book-review": _("book-review"),
                            "journal-issue": _("journal-issue"),
                            "review": _("review"),
                            "transcription": _("transcription"),
                            "translation": _("translation"),
                            "blog-post": _("blog-post"),
                            "dictionary-entry": _("dictionary-entry"),
                            "encyclopedia-entry": _("encyclopedia-entry"),
                            "magazine-article": _("magazine-article"),
                            "newspaper-article": _("newspaper-article"),
                            "newsletter-article": _("newsletter-article"),
                            "report": _("report"),
                            "public-speech": _("public-speech"),
                            "website": _("website"),
                            "artistic-performance": _("artistic-performance"),
                            "design": _("design"),
                            "image": _("image"),
                            "online-resource": _("online-resource"),
                            "moving-image": _("moving-image"),
                            "musical-composition": _("musical-composition"),
                            "sound": _("sound"),
                            "cartographic-material": _("cartographic-material"),
                            "clinical-study": _("clinical-study"),
                            "data-set": _("data-set"),
                            "data-management-plan": _("data-management-plan"),
                            "physical-object": _("physical-object"),
                            "research-technique": _("research-technique"),
                            "research-tool": _("research-tool"),
                            "software": _("software"),
                            "invention": _("invention"),
                            "licence": _("licence"),
                            "patent": _("patent"),
                            "registered-copyright": _("registered-copyright"),
                            "standards-and-policy": _("standards-and-policy"),
                            "trademark": _("trademark"),
                            "lecture-speech": _("lecture-speech"),
                            "learning-object": _("learning-object"),
                            "supervised-student-publication": _("supervised-student-publication"),
                            "manual": _("manual"),
                        }

def format_work_type_for_display(raw_type):
    if pd.isna(raw_type):
        return raw_type

    raw_value = str(raw_type).strip()
    if raw_value in type_labels:
        return type_labels[raw_value]
    else:
        return raw_value

with st.sidebar:
    
    st.header(":toolbox: " + _("app-title"))

    # Compact language chooser with clickable emojis.
    if "locale_picker" not in st.session_state:
        st.session_state.locale_picker = st.session_state.locale

    st.radio(
        "lang",
        options=["fr", "en"],
        format_func=lambda option: "🇫🇷" if option == "fr" else "🇬🇧",
        key="locale_picker",
        horizontal=True,
        label_visibility="collapsed",
    )

    if "orcid_list" in st.session_state:
        st.button(_("Réinitialiser"), type="secondary", on_click=reset_session_state)

    st.image("img/oiseau-orcidee.png")

    if overton_enabled and not overton_key:
        with st.expander(_("Clés API"), icon=":material/key:"):
            overton_key = st.text_input(_("Clé API Overton"), help=_("Une clé est nécessaire pour activer le lien direct vers Overton. Vous trouverez la vôtre dans les paramètres de votre compte Overton."))

    with st.expander(_("À propos"), icon=":material/help:"):
        st.image("img/BIBL-logo.png", link="https://www.bibl.ulaval.ca/services/soutien-a-ledition-savante-et-a-la-recherche/identifiants-uniques-perennes-orcid-doi-isbn-ror")
        st.markdown(_("about_text"))


if st.query_params and "tab" in st.query_params and st.query_params["tab"] in ["works", "activites", "resume", "suggestions"]:
    match st.query_params["tab"]:
        case "activites":
            default_tab = _("Autres activités")
        case "resume":
            default_tab = _("Résumé")
        case "suggestions":
            default_tab = _("Suggestions")
        case "works":
            default_tab = _("Travaux")
else:
    default_tab = _("Résumé")

tab_summary, tab_works, tab_compare, tab_suggest = st.tabs([_("Résumé"), _("Travaux"), _("Comparateur"), _("Suggestions")], default=default_tab)
    
# Check for ORCID from query params first and validate immediately
if "orcid_list" not in st.session_state:
    if st.query_params and "orcid" in st.query_params and st.query_params["orcid"]:
        # Parse from URL parameter
        orcid_from_url = st.query_params["orcid"]
        if isinstance(orcid_from_url, str):
            orcid_list = [orcid.strip() for orcid in orcid_from_url.split(',') if orcid.strip()]
        else:
            orcid_list = [str(orcid_from_url).strip()]
        
        # ORCID validation
        orcid_pattern = r'^[0-9a-zA-Z]{4}-[0-9a-zA-Z]{4}-[0-9a-zA-Z]{4}-[0-9a-zA-Z]{4}$'
        invalid_orcids = [orcid for orcid in orcid_list if not re.match(orcid_pattern, orcid)]
        
        if invalid_orcids:
            st.error(_("Format d'ORCID incorrect pour: {invalid_orcids}. Le format doit être XXXX-XXXX-XXXX-XXXX.").format(invalid_orcids=', '.join(invalid_orcids)))
            st.stop()
        
        # Store validated ORCID list from URL
        st.session_state.orcid_list = orcid_list
    else:

        col_input, col_file = st.columns(2)

        with col_input:
            orcid_input = st.text_input(_("Renseignez un numéro ORCID (ou séparez plusieurs ORCIDs par des virgules):"), key="orcid_input_field")

        with col_file:
            orcid_file = st.file_uploader(_("Ou téléversez un fichier (format texte, ORCIDs séparés par des virgules ou un par ligne):"), type=["txt"], key="orcid_file_upload")
        
        # Process file if uploaded
        orcid_list_from_file = []
        if orcid_file:
            file_content = orcid_file.read().decode("utf-8")
            # Parse by newlines and commas, removing comments
            for line in file_content.split('\n'):
                # Remove comments prefaced by #
                if '#' in line:
                    line = line.split('#')[0]
                # Now parse the remaining content
                for orcid in line.split(','):
                    cleaned = orcid.strip()
                    if cleaned:
                        orcid_list_from_file.append(cleaned)
        
        # Validate on button click OR when input exists (Enter key pressed) OR when file is uploaded
        if (st.button(_("Valider"), type="primary") or orcid_input or orcid_file) and (orcid_input or orcid_file):
            # Parse and normalize orcid_input to always be a list
            if orcid_input:
                if isinstance(orcid_input, str):
                    orcid_list = [orcid.strip() for orcid in orcid_input.split(',') if orcid.strip()]
                elif isinstance(orcid_input, list):
                    orcid_list = [orcid.strip() for orcid in orcid_input if orcid.strip()]
                else:
                    orcid_list = [str(orcid_input).strip()]
            else:
                orcid_list = []
            
            # Merge with file input
            orcid_list.extend(orcid_list_from_file)
            
            # Remove duplicates while preserving order
            seen = set()
            orcid_list = [x for x in orcid_list if not (x in seen or seen.add(x))]
            
            if not orcid_list:
                st.error(_("Veuillez fournir au moins un ORCID valide."))
                st.stop()
            
            # ORCID validation before storing
            orcid_pattern = r'^[0-9a-zA-Z]{4}-[0-9a-zA-Z]{4}-[0-9a-zA-Z]{4}-[0-9a-zA-Z]{4}$'
            invalid_orcids = [orcid for orcid in orcid_list if not re.match(orcid_pattern, orcid)]
            
            if invalid_orcids:
                st.error(_("Format d'ORCID incorrect pour: {invalid_orcids}. Le format doit être XXXX-XXXX-XXXX-XXXX.").format(invalid_orcids=', '.join(invalid_orcids)))
                st.stop()
            
            # Store in session state once validated
            st.session_state.orcid_list = orcid_list
            st.rerun()
        
        st.stop()

# Retrieve from session state
orcid_list = st.session_state.orcid_list

# Initialize storage for ORCID data if not exists
if 'orcid_data' not in st.session_state:
    st.session_state.orcid_data = {}

# Process each ORCID and store data
progress_text = _("Récupération des données ORCID...")
multifile_progress = st.progress(0, text=progress_text)
for idx, orcid_input in enumerate(orcid_list):     
    # Skip if already loaded
    if orcid_input not in st.session_state.orcid_data:
        with st.spinner(_("Chargement de {orcid_input}...").format(orcid_input=orcid_input)):
            try:
                df, raw, orcid_output, person_name = fetch_orcid_data(orcid_input)
            except Exception as e:
                st.error(_("Erreur lors de la récupération des données ORCID pour {orcid_input}: {error}").format(orcid_input=orcid_input, error=str(e)))
                continue
            works_count = len(df)

            summary_works = {
                "count": works_count,
                "last_modified": format_timestamp(raw.get('activities-summary', {}).get('works', {}). get('last-modified-date', {}).get('value'),False),
                "last_modified_display": format_timestamp(raw.get('activities-summary', {}).get('works', {}). get('last-modified-date', {}).get('value'),True),
                } if raw.get('activities-summary', {}).get('works', {}).get('last-modified-date') else None
        
            summary_employments = {
                "count": raw.get('activities-summary', {}).get('employments').get('affiliation-group', []).__len__(),
                "last_modified": format_timestamp(raw.get('activities-summary', {}).get('employments', {}). get('last-modified-date', {}).get('value'))
                } if raw.get('activities-summary', {}).get('employments', {}).get('last-modified-date') else None
            
            summary_educations = {
                "count": raw.get('activities-summary', {}).get('educations').get('affiliation-group', []).__len__(),
                "last_modified": format_timestamp(raw.get('activities-summary', {}).get('educations', {}). get('last-modified-date', {}).get('value'))
                } if raw.get('activities-summary', {}).get('educations', {}).get('last-modified-date') else None
            
            summary_fundings = {
                "count": raw.get('activities-summary', {}).get('fundings').get('affiliation-group', []).__len__(),
                "last_modified": format_timestamp(raw.get('activities-summary', {}).get('fundings', {}). get('last-modified-date', {}).get('value'),False),
                "last_modified_display": format_timestamp(raw.get('activities-summary', {}).get('fundings', {}). get('last-modified-date', {}).get('value'),True)
                } if raw.get('activities-summary', {}).get('fundings', {}).get('last-modified-date') else None
            
            try:
                updated_person = raw.get('person', {}).get('last-modified-date', {}).get('value')
            except Exception:
                updated_person = None

            # Store data in session state
            st.session_state.orcid_data[orcid_input] = {
                'df': df,
                'raw': raw,
                'person_name': person_name,
                'works_count': works_count,
                'summary_works': summary_works,
                'summary_employments': summary_employments,
                'summary_educations': summary_educations,
                'summary_fundings': summary_fundings,
                'updated_person': updated_person
            }
            multifile_progress.progress((idx + 1) / len(orcid_list), text=progress_text + f" ({idx + 1}/{len(orcid_list)})")
        
        # Show status message after loading ORCID data
        st.toast(_("Données ORCID chargées pour {orcid}.").format(orcid=orcid_list[0]), icon=":material/check_circle:")

multifile_progress.empty()
# For backward compatibility with single ORCID code
if len(orcid_list) == 1:
    orcid_input = orcid_list[0]
    df = st.session_state.orcid_data[orcid_input]['df']
    raw = st.session_state.orcid_data[orcid_input]['raw']
    person_name = st.session_state.orcid_data[orcid_input]['person_name']
    works_count = st.session_state.orcid_data[orcid_input]['works_count']
    summary_works = st.session_state.orcid_data[orcid_input]['summary_works']
    summary_employments = st.session_state.orcid_data[orcid_input]['summary_employments']
    summary_educations = st.session_state.orcid_data[orcid_input]['summary_educations']
    summary_fundings = st.session_state.orcid_data[orcid_input]['summary_fundings']
    updated_person = st.session_state.orcid_data[orcid_input]['updated_person']

# Create summary dataframe from all loaded ORCID data
orcid_summary_df = pd.DataFrame([
    {
        'orcid': orcid_id,
        'url': 'https://orcid.org/' + orcid_id,
        'person_name': data['person_name'],
        'person_last_modified': format_timestamp(data['updated_person']) if data['updated_person'] else None,
        'works_count': data['works_count'],
        'works_last_modified': data['summary_works']['last_modified'] if data['summary_works'] else None,
        'works_last_modified_display': data['summary_works']['last_modified_display'] if data['summary_works'] else None,
        'employments_count': data['summary_employments']['count'] if data['summary_employments'] else 0,
        'employments_last_modified': data['summary_employments']['last_modified'] if data['summary_employments'] else None,
        'educations_count': data['summary_educations']['count'] if data['summary_educations'] else 0,
        'educations_last_modified': data['summary_educations']['last_modified'] if data['summary_educations'] else None,
        'fundings_count': data['summary_fundings']['count'] if data['summary_fundings'] else 0,
        'fundings_last_modified': data['summary_fundings']['last_modified'] if data['summary_fundings'] else None,
        'fundings_last_modified_display': data['summary_fundings']['last_modified_display'] if data['summary_fundings'] else None,
        'drilldown' : '?tab=works&orcid=' + orcid_id
    }
    for orcid_id, data in st.session_state.orcid_data.items()
])
    
with tab_works:
    if len(orcid_list) == 1:
        if works_count == 0:
            st.warning(_("Aucun travail trouvé pour {person_name} ({orcid_input}).").format(person_name=person_name, orcid_input=orcid_input))
        else:
            works_df = df.copy()
            col1, col2 = st.columns([4,1],vertical_alignment="bottom")
            with col1:
                st.header(_("Travaux de {person_name}").format(person_name=person_name))
            with col2:
                st.link_button(_("Voir profil {orcid_input}").format(orcid_input=orcid_input), raw.get('orcid-identifier', {}).get('uri'), icon=":material/open_in_new:")     
    else:
        # Combine works from multiple profiles with ORCID and name columns
        dfs_with_orcid = []
        for orcid_id, data in st.session_state.orcid_data.items():
            df_copy = data['df'].copy()
            df_copy['orcid'] = orcid_id
            df_copy['name'] = data['person_name']
            dfs_with_orcid.append(df_copy)
        works_df = pd.concat(dfs_with_orcid, ignore_index=True)
        works_count = len(works_df)
        st.header(_("Travaux combinés de {count} profils").format(count=len(orcid_list)))

    if works_count > 0:
        with st.expander(":material/filter_alt: " + _("Filtrer")):
            filtered_df = works_df

            filter_col1, filter_col2 = st.columns([5, 1])
            
            with filter_col1:
                # Add an option to filter by type
                if 'type' in works_df.columns:
                    types = sorted(works_df['type'].dropna().unique().tolist())
                    if types:
                        selected_types = st.multiselect(
                            _("Filtrer par type:"),
                            types,
                            format_func=format_work_type_for_display,
                            placeholder=_("Sélectionnez les types de travaux à afficher")
                            )
                        if selected_types:
                            filtered_df = works_df[works_df['type'].isin(selected_types)]
                
                # Add an option to filter by publication year
                if 'publication-year' in works_df.columns:
                    years = sorted(works_df['publication-year'].dropna().unique().tolist())
                    if years:
                        lowest_year, highest_year = st.select_slider(
                            _("Filtrer par année de publication:"),
                            years,
                            value=(years[0], years[-1]),
                            )
                        if (lowest_year, highest_year) != (years[0], years[-1]):
                            filtered_df = filtered_df[(filtered_df['publication-year'] >= lowest_year) & (filtered_df['publication-year'] <= highest_year)]
                
                # Add an option to filter by Author if multiple profiles
                if len(orcid_list) > 1 and 'name' in works_df.columns:
                    names = sorted(works_df['name'].dropna().unique().tolist())
                    selected_names = st.multiselect(
                        _("Filtrer par chercheur:"),
                        names,
                        placeholder=_("Sélectionnez les chercheurs à afficher")
                        )
                    if selected_names:
                        filtered_df = filtered_df[filtered_df['name'].isin(selected_names)]
            
            with filter_col2:
                st.metric(_("Travaux affichés"), len(filtered_df), delta=f"{len(filtered_df) - len(works_df)} " + _("filtrés"))
                works_without_year = filtered_df['publication-year'].isna().sum()
                if works_without_year == 1:
                    st.badge(_("1 travail sans année de publication"), icon=":material/warning:", color="orange")
                elif works_without_year > 1:
                    st.badge(_("{count} travaux sans année de publication").format(count=works_without_year), icon=":material/warning:", color="orange")
        
        with st.expander(":material/export_notes: " + _("Exporter")):
            export_files_col, citation_col, export_overton_col = st.columns(3)

            with export_files_col:

                def prepare_works_for_export(df):
                    df_copy = df.copy()
                    
                    # Humanize type values if the column exists
                    if 'type' in df_copy.columns:

                        def humanize_type(value):
                            if pd.isna(value):
                                return value
                            raw_value = str(value).strip()
                            if raw_value in type_labels:
                                return type_labels[raw_value]

                            return " ".join(
                                word.capitalize()
                                for word in raw_value.replace("_", " ").replace("-", " ").split()
                            )
                        
                        df_copy['type'] = df_copy['type'].map(humanize_type)

                    # Remove columns that are not useful in the export and may contain complex nested data
                    df_copy.drop(columns=['visibility','external-ids', 'modified-date-display'], inplace=True, errors='ignore')

                    # Humanize modified date for export
                    if 'modified-date' in df_copy.columns:
                        df_copy['modified-date'] = df_copy['modified-date'].map(format_timestamp)

                    # Humanize column names
                    df_copy.rename(columns={
                        'put-code': _("Put-code"),
                        'modified-date': _("Dernière modification"),
                        'modified-by': _("Modifié par"),
                        'title': _("Titre"),
                        'type': _("Type"),
                        'journal-title': _("Titre de revue"),
                        'publication-year': _("Année"),
                        'url': _("URL"),
                        'doi': _("DOI"),
                        'isbn': _("ISBN"),
                        'orcid': _("ORCID") if len(orcid_list) > 1 else None,
                        'name': _("Chercheur") if len(orcid_list) > 1 else None
                    }, inplace=True)    
                    return df_copy

                def works_make_csv():
                    works_df_copy = prepare_works_for_export(filtered_df)
                    return works_df_copy.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label=_("Télécharger CSV"),
                    data=works_make_csv,
                    file_name=_("liste-travaux") + '.csv',
                    mime='text/csv',
                    key="download_csv",
                    icon=":material/download:"
                )

                def works_make_excel():
                    excel_buffer = BytesIO()
                    works_df_copy = prepare_works_for_export(filtered_df)
                    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                        works_df_copy.to_excel(writer, index=False, sheet_name=_("Travaux"))

                        ws = writer.sheets[_("Travaux")]

                        # Adjust column widths
                        for column_cells in ws.columns:
                            max_length = 0
                            column_letter = get_column_letter(column_cells[0].column)
                            for cell in column_cells:
                                try:
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                                except Exception:
                                    pass
                            adjusted_width = min((max_length + 2), 50)
                            ws.column_dimensions[column_letter].width = adjusted_width

                    excel_buffer.seek(0)
                    return excel_buffer.getvalue()
            
                st.download_button(
                    label=_("Télécharger vers Excel"),
                    data=works_make_excel,
                    file_name=_("liste-travaux") + '.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key="download_excel",
                    icon=":material/table_view:"
                )
                
            with citation_col:
                # List of supported CSL styles for export
                csl_styles = {
                    "apa": _("APA"),
                    "mla": _("MLA"),
                    "pour-reussir-note": _("Dionne"),
                    "lluelles": _("LLuelles"),
                    "mcgill": _("McGill"),
                    "vancouver-nlm": _("Vancouver - NLM"),
                    "chicago-author-date": _("Chicago"),
                    "ieee": _("IEEE"),
                }
                csl_style_for_export = st.selectbox(_("Style de citation"), options=csl_styles.keys(), format_func=lambda key: csl_styles[key], help=_("Style de citation utilisé pour formater les références exportées."))
                def works_make_citations(csl_format="apa", csl_locale=default_locale):
                    # Handle special styles
                    if csl_format == "mcgill" and csl_locale == "fr":
                        csl_format = "mcgill-fr"
                    elif csl_format == "mcgill" and csl_locale == "en":
                        csl_format = "mcgill-en"
                    citations_df = get_citations(filtered_df, csl_format=csl_format, csl_locale=csl_locale)
                    # Remove rows with errors and keep only the citation text for export
                    citation_lines = citations_df[citations_df['citation_error'].isna()]['citation'].dropna().astype(str).str.strip().tolist()
                    numbered_citation_lines = []
                    for row_num, citation in enumerate(citation_lines, start=1):
                        renumbered = re.sub(r"^\[\s*1\s*\]\s*", f"[{row_num}] ", citation)
                        renumbered = re.sub(r"^\s*1\.\s*", f"{row_num}. ", renumbered)
                        numbered_citation_lines.append(renumbered)
                    return "\n\n".join(numbered_citation_lines).encode('utf-8')
                
                if "locale" in st.session_state:
                    csl_locale = st.session_state.locale
                else:
                    csl_locale = default_locale
                st.download_button(
                    label=_("Télécharger les citations ({style})").format(style=csl_styles[csl_style_for_export]),
                    data=lambda: works_make_citations(csl_format=csl_style_for_export, csl_locale=csl_locale),
                    file_name=_("citations") + '.txt',
                    mime='text/plain',
                    key="download_citations",
                    icon=":material/article_person:"
                )
                st.badge(_("La qualité des citations est limitée."), icon=":material/warning:", color="orange")
            
            with export_overton_col:
                if overton_enabled and len(overton_key.strip()) > 0:
                    doi_list_for_overton = filtered_df['doi'].dropna().unique().tolist()
                    works_without_doi = filtered_df['doi'].isna().sum()
                    max_doi_count = 25000
                    if len(doi_list_for_overton) > max_doi_count:
                        st.error(_("Le set Overton est limité à {max_count} DOIs. La sélection actuelle en contient {current_count}.").format(max_count=max_doi_count, current_count=len(doi_list_for_overton)))
                    elif len(doi_list_for_overton) > 0:
                        if "overton_url" not in st.session_state:
                            st.session_state.overton_url = None
                        if "overton_last_generated_signature" not in st.session_state:
                            st.session_state.overton_last_generated_signature = None

                        current_doi_signature = tuple(sorted(doi_list_for_overton))
                        has_generated_url = bool(st.session_state.overton_url)
                        doi_list_has_changed = st.session_state.overton_last_generated_signature != current_doi_signature
                        should_enable_generate = (not has_generated_url) or doi_list_has_changed

                        if should_enable_generate:
                            if st.button(_("Générer le set Overton"), key="generate_overton_set", disabled=not should_enable_generate, icon=":material/list_alt_add:"):
                                try:
                                    st.session_state.overton_url = get_overton_set_url(doi_list_for_overton, overton_key)
                                    st.session_state.overton_last_generated_signature = current_doi_signature
                                    st.rerun()
                                except Exception as e:
                                    st.error(_("Erreur lors de la génération du set Overton: {error}").format(error=str(e)))
                        else:
                            st.link_button(_("Lancer la requête dans Overton"), st.session_state.overton_url, icon=":material/feature_search:")
                    if works_without_doi == 1:
                        st.badge(_("1 travail sans DOI ne sera pas inclus dans la requête Overton"), icon=":material/warning:", color="orange")
                    elif works_without_doi > 1:
                        st.badge(_("{count} travaux sans DOI ne seront pas inclus dans la requête Overton").format(count=works_without_doi), icon=":material/warning:", color="orange")
                elif overton_enabled:
                    st.warning(_("Renseignez une clé API dans l'onglet gauche pour activer l'export vers Overton."))

        # Show a simple table of works
        if len(orcid_list) == 1:
             work_display_columns = ["title", "journal-title", "publication-year", "type", "doi", "url"]
        else:
            work_display_columns = ["name", "title", "journal-title", "publication-year", "type", "doi", "url"]

        try:
            display_df = filtered_df.copy()
            if 'type' in display_df.columns:
                display_df['type'] = display_df['type'].map(format_work_type_for_display)

            st.dataframe(display_df,
                            column_config={
                                "put-code": None,
                                "modified-date": None,
                                "modified-by": None,
                                "title": _("Titre"),
                                "type": _("Type"),
                                "journal-title": _("Titre de revue"),
                                "publication-year": _("Année"),
                                "external-ids": None,
                                "visibility": None,
                                "doi": _("DOI"),
                                "isbn": None,
                                "url": st.column_config.LinkColumn(_("Lien"), display_text=":material/open_in_new:"),
                                "orcid": None if len(orcid_list) == 1 else _("ORCID"),
                                "name": None if len(orcid_list) == 1 else _("Chercheur")
                                },
                                column_order=work_display_columns, 
                            height="content", 
                            hide_index=True)
        except Exception:
            st.write(_("Aucun travail disponible à afficher."))

with tab_summary:

    if len(orcid_list) == 1:

        try:
            col1, col2 = st.columns([4,1],vertical_alignment="bottom")
            with col1:
                st.header(_("Résumé du profil ORCID de {person_name}").format(person_name=person_name))
            with col2:
                st.link_button(_("Voir profil {orcid_input}").format(orcid_input=orcid_input), raw.get('orcid-identifier', {}).get('uri'), icon=":material/open_in_new:")

            st.write(_("Créé le: {creation_date}").format(creation_date=format_timestamp(raw.get('history', {}).get('submission-date', {}).get('value'))))

            updated_table = {
                _("Section"): [
                    ":material/person: " + _("Informations personnelles"),
                    ":material/work: " + _("Emploi"),
                    ":material/school: " + _("Formation et qualifications"),
                    ":material/money: " + _("Financements"),
                    ":material/docs: " + _("Travaux")
                ],
                _("Complété"): [
                    "✅" if raw.get('person', {}).get('name') else "❌",
                    f"✅ ({summary_employments['count']})" if summary_employments else "❌",
                    f"✅ ({summary_educations['count']})" if summary_educations else "❌",
                    f"✅ ({summary_fundings['count']})" if summary_fundings else "❌",
                    f"✅ ({summary_works['count']})" if summary_works else "❌"
                ],
                _("Dernière modification"): [
                    format_timestamp(updated_person) if updated_person else "N/A",
                    summary_employments['last_modified'] if summary_employments else "N/A",
                    summary_educations['last_modified'] if summary_educations else "N/A",
                    summary_fundings['last_modified_display'] if summary_fundings else "N/A",
                    summary_works['last_modified_display'] if summary_works else "N/A"
                ]
            }

            # If works have not been modified in a while, add a recommendation
            try:
                if format_timestamp(raw.get('activities-summary', {}).get('works', {}). get('last-modified-date', {}).get('value'),True, True)[1] != "fresh":
                    with tab_suggest:
                        st.info(_("Votre section Travaux n'a pas été mise à jour depuis le {last_modified}. Pensez à ajouter ou mettre à jour vos publications pour refléter vos travaux récents.").format(last_modified=summary_works['last_modified']))
            except Exception:
                pass

            st.table(updated_table, border="horizontal")

        except Exception as e:
            st.error(_("Erreur lors de l'affichage du résumé: {error}").format(error=str(e)))
            import traceback
            st.code(traceback.format_exc())

    else:

        with st.expander(":material/export_notes: " + _("Exporter")):
            csv_col, xls_col = st.columns(2)

            def prepare_summary_for_export(df):
                df_copy = df.drop(columns=['drilldown','url','works_last_modified_display','fundings_last_modified_display'], inplace=False)
                df_copy.rename(columns={
                    'orcid': _("ORCID"),
                    'person_name': _("Nom"),
                    'person_last_modified': _("Màj profil"),
                    'works_count': _("Travaux"),
                    'works_last_modified': _("Màj travaux"),
                    'employments_count': _("Emplois"),
                    'employments_last_modified': _("Màj emplois"),
                    'educations_count': _("Formations"),
                    'educations_last_modified': _("Màj formations"),
                    'fundings_count': _("Financements"),
                    'fundings_last_modified': _("Màj financements")
                }, inplace=True)
                return df_copy

            with csv_col:
                def summary_make_csv():
                    orcid_summary_df_copy = prepare_summary_for_export(orcid_summary_df)
                    return orcid_summary_df_copy.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label=_("Télécharger CSV"),
                    data=summary_make_csv,
                    file_name=_("resume-orcid") + '.csv',
                    mime='text/csv',
                    key="summary_download_csv",
                    icon=":material/download:"
                )

            with xls_col:
                def summary_make_excel():
                        excel_buffer = BytesIO()
                        orcid_summary_df_copy = prepare_summary_for_export(orcid_summary_df)
                        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                            orcid_summary_df_copy.to_excel(writer, index=False, sheet_name=_("Résumé"))
                            ws = writer.sheets[_("Résumé")]

                            last_row = ws.max_row

                            # Adjust column widths
                            for column_cells in ws.columns:
                                max_length = 0
                                column_letter = get_column_letter(column_cells[0].column)
                                for cell in column_cells:
                                    try:
                                        cell_length = len(str(cell.value))
                                        if cell_length > max_length:
                                            max_length = cell_length
                                    except Exception:
                                        pass
                                adjusted_width = (max_length + 2) * 1.2
                                ws.column_dimensions[column_letter].width = adjusted_width

                            # Apply conditional formatting for last modified dates
                            date_columns = {_('Màj travaux'),_("Màj financements")}
                            for col_name in date_columns:
                                column_letter = get_column_letter(orcid_summary_df_copy.columns.get_loc(col_name) + 1)
                                stale_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                aging_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                fresh_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                                date_range = f"{column_letter}2:{column_letter}{last_row}"

                                ws.conditional_formatting.add(
                                    date_range,
                                    FormulaRule(
                                        formula=[f'AND({column_letter}2<>"",TODAY()-DATEVALUE({column_letter}2)>730)'],
                                        fill=stale_fill
                                    )
                                )
                                ws.conditional_formatting.add(
                                    date_range,
                                    FormulaRule(
                                        formula=[f'AND({column_letter}2<>"",TODAY()-DATEVALUE({column_letter}2)>365,TODAY()-DATEVALUE({column_letter}2)<=730)'],
                                        fill=aging_fill
                                    )
                                )
                                ws.conditional_formatting.add(
                                    date_range,
                                    FormulaRule(
                                        formula=[f'AND({column_letter}2<>"",TODAY()-DATEVALUE({column_letter}2)<=365)'],
                                        fill=fresh_fill
                                    )
                                )

                        excel_buffer.seek(0)
                        return excel_buffer.getvalue()
            
                st.download_button(
                    label=_("Télécharger vers Excel"),
                    data=summary_make_excel,
                    file_name=_("resume-orcid") + '.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key="summary_download_excel",
                    icon=":material/table_view:"
                )


        st.dataframe(orcid_summary_df, column_config={
            "orcid": None,
            "url": st.column_config.LinkColumn(_("ORCID"), display_text="https://orcid.org/(.*)"),
            "drilldown": st.column_config.LinkColumn(_("Ouvrir détails"), display_text=":material/open_in_new:"),
            "person_name": _("Nom"),
            "works_count": _("Travaux"),
            "works_last_modified": None,
            "works_last_modified_display": _("Màj travaux"),
            "employments_count": None,
            "employments_last_modified": _("Màj emplois"),
            "educations_count": None,
            "educations_last_modified": _("Màj formations"),
            "fundings_count": None,
            "fundings_last_modified": None,
            "fundings_last_modified_display": _("Màj financements"),
            "person_last_modified": _("Màj profil")
            },
            column_order=[
                "url", "person_name","person_last_modified","works_count","works_last_modified_display","drilldown","employment_last_modified",
                "educations_last_modified","fundings_last_modified_display"],
            height="content",
            hide_index=True)
        
    st.subheader(_("Distribution des travaux par année de publication"))

    if works_count > 0 and 'publication-year' in works_df.columns:
        st.bar_chart(works_df['publication-year'].value_counts().sort_index())
    else:
        st.warning(_("Aucune donnée de publication disponible pour générer le graphique."))

    st.subheader(_("Distribution des travaux par type"))

    if works_count > 0 and 'type' in works_df.columns:
        mapped_work_types = works_df['type'].map(format_work_type_for_display)
        st.bar_chart(mapped_work_types.value_counts(), horizontal=True, sort=False)
    else:
        st.warning(_("Aucune donnée de publication disponible pour générer le graphique."))

with tab_suggest:
    st.warning(_("Cette section n'est pas encore implémentée."))

with tab_compare:

    if len(orcid_list) > 1:
        st.warning(_("Le comparateur ne peut être utilisé qu'avec un seul ORCID à la fois. Utilisez l'onglet 'Résumé' pour voir les données agrégées."))
        st.stop()

    if works_count == 0:
        st.warning(_("Aucun travail trouvé pour {person_name} ({orcid_input}). Le comparateur nécessite des travaux pour fonctionner.").format(person_name=person_name, orcid_input=orcid_input))
        st.stop()
    
    if importlib.util.find_spec("transformers") is None and importlib.util.find_spec("references_tractor") is None:
        st.warning(_("Cette fonctionalité nécessite la présence d'une bibliothèque pour l'extraction des références, telle que 'transformers' ou 'references_tractor'. Veuillez installer au moins l'une de ces bibliothèques."))
        st.stop()

    col_file, col_controls = st.columns(2)

    with col_file:

        refs_file = st.file_uploader(_("Téléchargez un fichier texte contenant des références bibliographiques à extraire :"), type=["txt"])
        
        # Initialize variables
        matched_refs = []
        unmatched_refs = []
        
        if refs_file:
            source_refs = refs_file.read().decode("utf-8")
            
            # Create progress bar for reference extraction
            extraction_progress = st.progress(0, text=_("Extraction des références en cours..."))
            
            def update_progress(current, total):
                progress_value = current / total if total > 0 else 0
                extraction_progress.progress(progress_value, text=_("Traitement des références... ({current}/{total})").format(current=current, total=total))
            
            # Extract and process references
            screened_refs, invalid_refs = extract_and_process_references(source_refs, progress_callback=update_progress)
            
            # Clear progress bar when done
            extraction_progress.empty()

            
            st.toast(_("{valid} références valides extraites, {invalid} références invalides ignorées.").format(valid=len(screened_refs), invalid=len(invalid_refs)), icon=":material/check_circle:")

    with col_controls:
        
        if refs_file:
            # Compare references with fuzzy matching
            st.markdown(_("**Contrôle de correspondance :**"))
            
            # Configure matching thresholds
            confidence_interval = st.slider(_("Seuil de confiance (%)"), 50, 100, (60, 90), 1)
            
            # Prepare ORCID works and match references
            orcid_works = prepare_orcid_works(df)
            matched_refs, unmatched_refs = match_references_to_orcid(screened_refs, orcid_works, confidence_interval[1])
            
            # Display statistics
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.metric(_("Références extraites"), len(screened_refs))
            with col_b:
                st.metric(_("Trouvées dans ORCID"), len(matched_refs))
            with col_c:
                st.metric(_("Manquantes dans ORCID"), len(unmatched_refs))
                

    if matched_refs:
        st.subheader("✅ " + _("{count} références trouvées dans ORCID").format(count=len(matched_refs)))
        sorting_option = st.segmented_control(_("Trier par :"), [_("Score"), _("Alpha"), _("Ordre")], key="sorting_option")
        for ref in matched_refs:
            col_source, col_target = st.columns(2)
            with col_source:
                ref_number = ref['ref_number']
                ref_ner = ref['ref_ner']
                ref_title_display = ref_ner["TITLE"][0] if "TITLE" in ref_ner and ref_ner["TITLE"] else ref["text"][:50] + "..."
                with st.expander(f"[{ref_number}] {ref_title_display}"):
                    st.caption(_("Texte original:"))
                    st.write(ref.get('ref', {}).get('text', ''))
                    col_inner, col_outer = st.columns(2)
                    with col_inner:
                        if ref.get('ref_journal'):
                            st.caption(_("Journal") + " : " + (ref['ref_journal'] or 'N/A'))
                        if ref.get('ref_year'):
                            st.caption(_("Année") + " : " + (ref['ref_year'] or 'N/A'))
                        if ref.get('ref_doi'):
                            st.caption(_("DOI") + " : " + (ref['ref_doi'] or 'N/A'))
                    with col_outer:
                        st.caption(_("Entités détectées :"))
                        st.json(ref_ner, expanded=False)

            with col_target:
                confidence_color = "🟢" if ref['confidence'] >= 90 else "🟡" if ref['confidence'] >= 80 else "🟠"
                with st.expander(f"{confidence_color} {ref['confidence']:.0f}% - {ref['orcid_title']}"):
                    st.caption(_("Score titre") + f" : {ref['title_score']}")
                    if ref.get('orcid_journal'):
                        st.caption(_("Journal") + f" : {ref['orcid_journal'] or 'N/A'} (score {ref['journal_score']})")
                    if ref.get('orcid_year'):
                        st.caption(_("Année") + f" : {ref['orcid_year'] or 'N/A'} (score {ref['year_score']})")
                    if ref.get('orcid_doi'):
                        st.caption(_("DOI") + f" : {ref['orcid_doi'] or 'N/A'} (score {ref['doi_score']})")
    
    if unmatched_refs:
        st.subheader("⚠️ " + _("Références à valider"))

        # Sort by confidence descending
        unmatched_refs_sorted = sorted(unmatched_refs, key=lambda x: x['confidence'], reverse=True)

        for ref in unmatched_refs_sorted:
            if confidence_interval[0] <= ref['confidence'] <= confidence_interval[1]:
                col_source, col_target = st.columns(2)
                with col_source:
                    ref_number = ref['ref_number']
                    ref_ner = ref['ref_ner']
                    ref_title_display = ref_ner["TITLE"][0] if "TITLE" in ref_ner and ref_ner["TITLE"] else ref["text"][:50] + "..."
                    with st.expander(f"[{ref_number}] {ref_title_display}"):
                        st.caption(_("Texte original:"))
                        st.write(ref.get('ref', {}).get('text', ''))
                        col_inner, col_outer = st.columns(2)
                        with col_inner:
                            if ref.get('ref_journal'):
                                st.caption(_("Journal") + " : " + (ref['ref_journal'] or 'N/A'))
                            if ref.get('ref_year'):
                                st.caption(_("Année") + " : " + (ref['ref_year'] or 'N/A'))
                            if ref.get('ref_doi'):
                                st.caption(_("DOI") + " : " + (ref['ref_doi'] or 'N/A'))
                        with col_outer:
                            st.caption(_("Entités détectées :"))
                            st.json(ref_ner, expanded=False)

                with col_target:
                    confidence_color = "🟢" if ref['confidence'] >= 90 else "🟡" if ref['confidence'] >= 80 else "🟠"
                    with st.expander(f"{confidence_color} {ref['confidence']:.0f}% - {ref['orcid_title']}"):
                        st.caption(_("Score titre") + f" : {ref['title_score']}")
                        if ref.get('orcid_journal'):
                            st.caption(_("Journal: {journal} (score {score})").format(journal=ref['orcid_journal'] or 'N/A', score=ref['journal_score']))
                        if ref.get('orcid_year'):
                            st.caption(_("Année: {year} (score {score})").format(year=ref['orcid_year'] or 'N/A', score=ref['year_score']))
                        if ref.get('orcid_doi'):    
                            st.caption(_("DOI: {doi} (score {score})").format(doi=ref['orcid_doi'] or 'N/A', score=ref['doi_score']))
        
        st.subheader("❌ " + _("Références non trouvées"))

        for ref in unmatched_refs_sorted:
            if confidence_interval[0] > ref['confidence'] :
                col_source, col_target = st.columns(2)
                with col_source:
                    ref_number = ref['ref_number']
                    ref_ner = ref['ref_ner']
                    ref_title_display = ref_ner["TITLE"][0] if "TITLE" in ref_ner and ref_ner["TITLE"] else ref["text"][:50] + "..."
                    with st.expander(f"[{ref_number}] {ref_title_display}"):
                        st.write(ref.get('ref', {}).get('text', ''))
                        col_inner, col_outer = st.columns(2)
                        with col_inner:
                            if ref.get('ref_journal'):
                                st.caption(_("Journal") + " : " + (ref['ref_journal'] or 'N/A'))
                            if ref.get('ref_year'):
                                st.caption(_("Année") + " : " + (ref['ref_year'] or 'N/A'))
                            if ref.get('ref_doi'):
                                st.caption(_("DOI") + " : " + (ref['ref_doi'] or 'N/A'))
                        with col_outer:
                            st.caption(_("Entités détectées :"))
                            st.json(ref_ner, expanded=False)
            


